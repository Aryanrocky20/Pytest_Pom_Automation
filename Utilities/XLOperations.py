import openpyxl


class ReadExcel:

    def __init__(self, fileName, sheetName):
        self.fileName = fileName
        self.sheetName = sheetName
        self.workbook = openpyxl.load_workbook(filename=self.fileName)
        self.sheet = self.workbook.get_sheet_by_name(self.sheetName)

    def readRowCount(self):
        return self.sheet.max_row

    def readColumnCount(self):
        return self.sheet.max_column

    def readData(self, row, column):
        return self.sheet.cell(row=row, column=column)

    def writeData(self, row, column, data):
        self.sheet.cell(row=row, column=column).value = data
        self.workbook.save(self.fileName)

